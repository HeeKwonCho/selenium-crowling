'''
언어 : 파이썬
라이브러리 : 셀레니움(selenium)
분야 : 크롤링
사이트 : https://pubmed.ncbi.nlm.nih.gov/?term=covid-19
방법 :
1. https://pubmed.ncbi.nlm.nih.gov/?term=covid-19 웹페이지를 로딩
2. 하나의 페이지에 full-docsum이 10개 씩 존재
3. 크롤링 할 것들 : full-docsum안의 docsum-title, docsum-authors.full-authors
docsum-journal-citation.full-journal-citation,citation-part(PMID),full-view-snippet
을 크롤링
4. 100개 에서 200개 정도 크롤링. (페이지로 따지면 10페이지에서 20페이지)
5. 저 자료들을 먼저 리스트로 받은다음 csv파일로 저장할 수 있는지?
6. 최종적으로 csv파일로 저장하는 과정까지
'''

from subprocess import CREATE_NO_WINDOW
from time import sleep

import chromedriver_autoinstaller
import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  # 크롬드라이버 버전 확인

try:
    s = Service(f'./{chrome_ver}/chromedriver.exe')
    s.creationflags = CREATE_NO_WINDOW
    driver = webdriver.Chrome(service=s)
except:
    chromedriver_autoinstaller.install(True)
    s = Service(f'./{chrome_ver}/chromedriver.exe')
    s.creationflags = CREATE_NO_WINDOW
    driver = webdriver.Chrome(service=s)
driver.implicitly_wait(4)

wb = openpyxl.Workbook()
ws = wb.active
i=1

# 엑셀 첫째줄 제목 적기
sub = ['제목', 'Authors', 'Citation', 'PMID', 'Full view']
for kwd, j in zip(sub, list(range(1, len(sub)+1))):
    ws.cell(row=1, column=j).value = kwd

ws.freeze_panes = 'A2' # 첫행고정

# 셀너비
ws.column_dimensions['A'].width = 135 # A열
ws.column_dimensions['B'].width = 123 # B열
ws.column_dimensions['C'].width = 97 # C열
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 145

# 셀 1행 배경색 노랑색 변경
y_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
for num in range(1, len(sub)+1):
    ws.cell(1,num).fill = y_color

i=2
driver.get('https://pubmed.ncbi.nlm.nih.gov/?term=covid-19')
sleep(1)

articleElements = driver.find_elements(By.CSS_SELECTOR,'article.full-docsum') # 리스트 목록 elements

pagesToCrawl = 200 # 200페이지 크롤링, 100을 쓰면 100페이지 크롤링
for _ in range(pagesToCrawl):
    for elem in articleElements:
        title = elem.find_element(By.CSS_SELECTOR,'a.docsum-title').text
        authors = elem.find_element(By.CSS_SELECTOR,'span.docsum-authors.full-authors').text
        citation = elem.find_element(By.CSS_SELECTOR,'span.docsum-journal-citation.full-journal-citation').text
        PMID = elem.find_element(By.CSS_SELECTOR,'span.docsum-pmid').text
        fullview = elem.find_element(By.CSS_SELECTOR,'div.full-view-snippet').text

        ws['A'+str(i)] = title
        ws['B'+str(i)] = authors
        ws['C'+str(i)] = citation
        ws['D'+str(i)] = PMID
        ws['E'+str(i)] = fullview
        i+=1
    driver.find_element(By.CSS_SELECTOR,'button.button-wrapper.next-page-btn').click() # Next 버튼 클릭
    sleep(1)
    articleElements = driver.find_elements(By.CSS_SELECTOR, 'article.full-docsum')  # 리스트 목록 elements
wb.save('크롤링결과.xlsx')
wb.close()
print(f'총{pagesToCrawl} 페이지를 크롤링하였습니다')
print(f'작업이 완료되었습니다.')
